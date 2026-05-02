"""
ANVISA/CMED Client - Integração com a tabela de preços da Secretaria Executiva da CMED.

Funcionalidades:
1. Download e importação da planilha CMED (27k+ medicamentos)
2. Autenticação OAuth2 com o Portal de APIs da ANVISA (quando disponível)
3. Busca fuzzy de medicamentos por substância/produto
4. Cruzamento de itens de licitação com preços CMED de referência
"""

import httpx
import logging
import os
import io
from datetime import datetime
from typing import List, Optional, Dict, Any
from sqlmodel.ext.asyncio.session import AsyncSession
from sqlmodel import select, col
from models import MedicamentoCMED

logger = logging.getLogger("AnvisaClient")


class AnvisaClient:
    """
    Client para integração com dados da ANVISA/CMED.

    Suporta dois modos:
    1. Download direto da planilha CMED (sem autenticação)
    2. API OAuth2 da ANVISA (requer client_id/client_secret)
    """

    # URL da planilha CMED mais recente (atualizada periodicamente pela ANVISA)
    CMED_XLS_URL = "https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos/lista-de-precos-de-medicamentos"
    
    # API Gateway ANVISA (requer OAuth)
    API_BASE = "https://api-gateway.prd.apps.anvisa.gov.br"
    TOKEN_URL = f"{API_BASE}/oauth/token"

    def __init__(self):
        self.client = httpx.AsyncClient(timeout=120.0, follow_redirects=True)
        self.access_token: Optional[str] = None
        
        # Credenciais OAuth (do .env)
        self.client_id = os.getenv("ANVISA_CLIENT_ID", "")
        self.client_secret = os.getenv("ANVISA_CLIENT_SECRET", "")

    async def close(self):
        await self.client.aclose()

    # ─── OAuth2 Authentication ────────────────────────────────────────────────

    async def authenticate(self) -> bool:
        """
        Obtém Bearer Token via OAuth2 client_credentials.
        Requer ANVISA_CLIENT_ID e ANVISA_CLIENT_SECRET no .env
        """
        if not self.client_id or not self.client_secret:
            logger.warning("⚠️ ANVISA OAuth: Credenciais não configuradas. Usando modo planilha.")
            return False

        try:
            response = await self.client.post(
                self.TOKEN_URL,
                data={
                    "grant_type": "client_credentials",
                    "client_id": self.client_id,
                    "client_secret": self.client_secret,
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"}
            )
            
            if response.status_code == 200:
                data = response.json()
                self.access_token = data.get("access_token")
                logger.info(f"✅ ANVISA OAuth: Token obtido. Expira em {data.get('expires_in', '?')}s")
                return True
            else:
                logger.error(f"❌ ANVISA OAuth falhou: {response.status_code} - {response.text[:200]}")
                return False
        except Exception as e:
            logger.error(f"❌ ANVISA OAuth erro: {e}")
            return False

    def _auth_headers(self) -> dict:
        """Retorna headers com Bearer Token se autenticado."""
        if self.access_token:
            return {"Authorization": f"Bearer {self.access_token}"}
        return {}

    # ─── Importação da Planilha CMED ──────────────────────────────────────────

    async def import_cmed_from_xls(self, session: AsyncSession, file_path: str = None, file_bytes: bytes = None) -> int:
        """
        Importa a planilha CMED (XLS/XLSX) para o banco de dados.
        
        Aceita:
        - file_path: caminho local para o arquivo XLS/XLSX
        - file_bytes: bytes do arquivo (upload via API)
        
        Retorna: número de medicamentos importados/atualizados.
        """
        try:
            import openpyxl
        except ImportError:
            logger.error("❌ openpyxl não instalado. Execute: pip install openpyxl")
            return 0

        count = 0
        
        try:
            if file_bytes:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            elif file_path:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            else:
                logger.error("❌ Nenhum arquivo fornecido.")
                return 0

            # A planilha CMED geralmente tem a aba principal como a primeira
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            
            if not rows:
                logger.error("❌ Planilha vazia.")
                return 0

            # Detectar linha de cabeçalho (procura por "SUBSTÂNCIA" ou "PRODUTO")
            header_row_idx = 0
            for i, row in enumerate(rows):
                row_str = " ".join(str(c or "").upper() for c in row)
                if "SUBSTÂNCIA" in row_str or "SUBSTANCIA" in row_str or "PRODUTO" in row_str:
                    header_row_idx = i
                    break

            headers = [str(h or "").strip().upper() for h in rows[header_row_idx]]
            logger.info(f"📋 Cabeçalho detectado na linha {header_row_idx + 1}: {headers[:10]}...")

            # Mapear colunas por nome
            col_map = {}
            for idx, h in enumerate(headers):
                col_map[h] = idx

            def get_val(row, *possible_names, default=None):
                for name in possible_names:
                    name_upper = name.upper()
                    if name_upper in col_map:
                        val = row[col_map[name_upper]]
                        return val if val is not None else default
                return default

            def safe_float(val) -> Optional[float]:
                if val is None:
                    return None
                try:
                    if isinstance(val, (int, float)):
                        return float(val)
                    val_str = str(val).replace(".", "").replace(",", ".").strip()
                    return float(val_str) if val_str else None
                except (ValueError, TypeError):
                    return None

            def safe_bool(val) -> bool:
                if val is None:
                    return False
                val_str = str(val).strip().upper()
                return val_str in ("SIM", "S", "TRUE", "1", "X", "YES")

            # Limpar dados antigos antes de importar
            await session.exec(select(MedicamentoCMED).where(MedicamentoCMED.id > 0))
            # Deletar todos os registros existentes
            from sqlalchemy import delete
            await session.exec(delete(MedicamentoCMED))
            await session.commit()
            logger.info("🗑️ Dados CMED antigos removidos.")

            # Processar linhas de dados (após cabeçalho)
            batch_size = 500
            batch = []
            
            for row in rows[header_row_idx + 1:]:
                # Skip linhas vazias
                if not row or all(c is None for c in row):
                    continue
                
                substancia = str(get_val(row, "SUBSTÂNCIA", "SUBSTANCIA", "PRINCÍPIO ATIVO", default="") or "").strip()
                produto = str(get_val(row, "PRODUTO", "MEDICAMENTO", "NOME COMERCIAL", default="") or "").strip()
                
                if not substancia and not produto:
                    continue

                med = MedicamentoCMED(
                    substancia=substancia,
                    cnpj=str(get_val(row, "CNPJ", default="") or "").strip(),
                    laboratorio=str(get_val(row, "LABORATÓRIO", "LABORATORIO", default="Desconhecido") or "Desconhecido").strip(),
                    codigo_ggrem=str(get_val(row, "CÓDIGO GGREM", "CODIGO GGREM", default="") or "").strip() or None,
                    registro=str(get_val(row, "REGISTRO", default="") or "").strip() or None,
                    ean_1=str(get_val(row, "EAN 1", "EAN1", default="") or "").strip() or None,
                    ean_2=str(get_val(row, "EAN 2", "EAN2", default="") or "").strip() or None,
                    produto=produto,
                    apresentacao=str(get_val(row, "APRESENTAÇÃO", "APRESENTACAO", default="") or "").strip(),
                    classe_terapeutica=str(get_val(row, "CLASSE TERAPÊUTICA", "CLASSE TERAPEUTICA", default="") or "").strip() or None,
                    tipo_produto=str(get_val(row, "TIPO DE PRODUTO (STATUS DO PRODUTO)", "TIPO PRODUTO", default="") or "").strip() or None,
                    regime_preco=str(get_val(row, "REGIME DE PREÇO", "REGIME DE PRECO", default="") or "").strip() or None,
                    
                    # Preços Fábrica
                    pf_sem_impostos=safe_float(get_val(row, "PF SEM IMPOSTOS", "PF S/ IMPOSTOS")),
                    pf_0=safe_float(get_val(row, "PF 0%")),
                    pf_12=safe_float(get_val(row, "PF 12%")),
                    pf_17=safe_float(get_val(row, "PF 17%")),
                    pf_17_alc=safe_float(get_val(row, "PF 17% ALC")),
                    pf_17_5=safe_float(get_val(row, "PF 17,5%")),
                    pf_18=safe_float(get_val(row, "PF 18%")),
                    pf_19=safe_float(get_val(row, "PF 19%")),
                    pf_20=safe_float(get_val(row, "PF 20%")),
                    pf_21=safe_float(get_val(row, "PF 21%")),
                    pf_22=safe_float(get_val(row, "PF 22%")),
                    
                    # Preços Máximos ao Consumidor
                    pmc_0=safe_float(get_val(row, "PMC 0%")),
                    pmc_12=safe_float(get_val(row, "PMC 12%")),
                    pmc_17=safe_float(get_val(row, "PMC 17%")),
                    pmc_17_alc=safe_float(get_val(row, "PMC 17% ALC")),
                    pmc_17_5=safe_float(get_val(row, "PMC 17,5%")),
                    pmc_18=safe_float(get_val(row, "PMC 18%")),
                    pmc_19=safe_float(get_val(row, "PMC 19%")),
                    pmc_20=safe_float(get_val(row, "PMC 20%")),
                    pmc_21=safe_float(get_val(row, "PMC 21%")),
                    pmc_22=safe_float(get_val(row, "PMC 22%")),
                    
                    # Flags
                    restricao_hospitalar=safe_bool(get_val(row, "RESTRIÇÃO HOSPITALAR", "RESTRICAO HOSPITALAR")),
                    cap=safe_bool(get_val(row, "CAP")),
                    confaz_87=safe_bool(get_val(row, "CONFAZ 87")),
                    icms_0=safe_bool(get_val(row, "ICMS 0%")),
                    analise_recurso=safe_bool(get_val(row, "ANÁLISE RECURSO", "ANALISE RECURSO")),
                    lista_concessao_credito_pis_cofins=str(get_val(row, "LISTA DE CONCESSÃO DE CRÉDITO TRIBUTÁRIO (PIS/COFINS)", default="") or "").strip() or None,
                    tarja=str(get_val(row, "TARJA", default="") or "").strip() or None,
                    
                    data_atualizacao=datetime.utcnow()
                )
                batch.append(med)
                count += 1

                if len(batch) >= batch_size:
                    session.add_all(batch)
                    await session.commit()
                    batch = []
                    logger.info(f"📦 Importados {count} medicamentos...")

            # Commit restante
            if batch:
                session.add_all(batch)
                await session.commit()

            wb.close()
            logger.info(f"✅ Importação CMED concluída: {count} medicamentos importados.")
            return count

        except Exception as e:
            logger.error(f"❌ Erro na importação CMED: {e}")
            import traceback
            traceback.print_exc()
            return 0

    # ─── Busca de Medicamentos ────────────────────────────────────────────────

    @staticmethod
    async def search_medicamentos(
        session: AsyncSession,
        query: str,
        limit: int = 20,
        tipo_produto: str = None,
        classe_terapeutica: str = None,
    ) -> List[MedicamentoCMED]:
        """
        Busca fuzzy de medicamentos na base CMED.
        Pesquisa por substância, produto ou apresentação.
        """
        search_term = f"%{query.upper()}%"
        
        stmt = select(MedicamentoCMED).where(
            (col(MedicamentoCMED.substancia).ilike(search_term)) |
            (col(MedicamentoCMED.produto).ilike(search_term)) |
            (col(MedicamentoCMED.apresentacao).ilike(search_term))
        )
        
        if tipo_produto:
            stmt = stmt.where(col(MedicamentoCMED.tipo_produto).ilike(f"%{tipo_produto}%"))
        if classe_terapeutica:
            stmt = stmt.where(col(MedicamentoCMED.classe_terapeutica).ilike(f"%{classe_terapeutica}%"))
        
        stmt = stmt.limit(limit)
        result = await session.exec(stmt)
        return list(result.all())

    # ─── Cruzamento com Itens de Licitação ────────────────────────────────────

    @staticmethod
    async def cruzar_itens_com_cmed(
        session: AsyncSession,
        itens_descricoes: List[str],
    ) -> List[Dict[str, Any]]:
        """
        Para cada descrição de item de uma licitação, busca o preço CMED de referência.
        Retorna lista de matches com preço mínimo e máximo encontrados.
        """
        resultados = []

        for desc in itens_descricoes:
            # Extrair palavras-chave relevantes (substância principal)
            desc_clean = desc.upper().strip()
            
            # Buscar match direto
            search_term = f"%{desc_clean[:60]}%"  # Limita para evitar queries enormes
            
            stmt = select(MedicamentoCMED).where(
                (col(MedicamentoCMED.substancia).ilike(search_term)) |
                (col(MedicamentoCMED.produto).ilike(search_term))
            ).limit(5)
            
            result = await session.exec(stmt)
            matches = list(result.all())

            if not matches:
                # Tentativa com palavras individuais (top 2 mais longas)
                palavras = sorted(desc_clean.split(), key=len, reverse=True)[:2]
                for p in palavras:
                    if len(p) < 4:
                        continue
                    stmt2 = select(MedicamentoCMED).where(
                        (col(MedicamentoCMED.substancia).ilike(f"%{p}%")) |
                        (col(MedicamentoCMED.produto).ilike(f"%{p}%"))
                    ).limit(5)
                    result2 = await session.exec(stmt2)
                    matches = list(result2.all())
                    if matches:
                        break

            if matches:
                precos_pf = [m.pf_18 or m.pf_17 or m.pf_0 or 0 for m in matches if (m.pf_18 or m.pf_17 or m.pf_0)]
                resultados.append({
                    "descricao_item": desc,
                    "match_encontrado": True,
                    "quantidade_matches": len(matches),
                    "melhor_match": {
                        "substancia": matches[0].substancia,
                        "produto": matches[0].produto,
                        "laboratorio": matches[0].laboratorio,
                        "apresentacao": matches[0].apresentacao,
                        "tipo": matches[0].tipo_produto,
                    },
                    "preco_fabrica_min": min(precos_pf) if precos_pf else None,
                    "preco_fabrica_max": max(precos_pf) if precos_pf else None,
                    "preco_fabrica_medio": sum(precos_pf) / len(precos_pf) if precos_pf else None,
                    "todos_matches": [
                        {
                            "substancia": m.substancia,
                            "produto": m.produto,
                            "lab": m.laboratorio,
                            "pf_18": m.pf_18,
                            "pf_17": m.pf_17,
                            "pf_0": m.pf_0,
                            "tipo": m.tipo_produto,
                        } for m in matches
                    ]
                })
            else:
                resultados.append({
                    "descricao_item": desc,
                    "match_encontrado": False,
                    "quantidade_matches": 0,
                })

        return resultados

    # ─── API ANVISA (Consultas Externas) ──────────────────────────────────────

    async def consultar_medicamento_api(self, nome_medicamento: str) -> Optional[Dict]:
        """
        Consulta a API REST da ANVISA para dados de medicamentos regulados.
        Requer autenticação OAuth2.
        """
        if not self.access_token:
            authenticated = await self.authenticate()
            if not authenticated:
                return None

        try:
            url = f"{self.API_BASE}/consultas-externas-api/api/v1/medicamentos"
            params = {"filter[nomeProduto]": nome_medicamento, "page[size]": 10}
            
            response = await self.client.get(url, params=params, headers=self._auth_headers())

            if response.status_code == 200:
                return response.json()
            elif response.status_code == 401:
                logger.warning("🔄 Token expirado. Re-autenticando...")
                if await self.authenticate():
                    response = await self.client.get(url, params=params, headers=self._auth_headers())
                    if response.status_code == 200:
                        return response.json()
            
            logger.warning(f"ANVISA API: {response.status_code}")
            return None

        except Exception as e:
            logger.error(f"Erro na API ANVISA: {e}")
            return None

    # ─── Stats ────────────────────────────────────────────────────────────────

    @staticmethod
    async def get_stats(session: AsyncSession) -> Dict[str, Any]:
        """Retorna estatísticas da base CMED local."""
        from sqlalchemy import func
        
        total = await session.exec(select(func.count(MedicamentoCMED.id)))
        total_count = total.one()

        tipos = await session.exec(
            select(MedicamentoCMED.tipo_produto, func.count(MedicamentoCMED.id))
            .group_by(MedicamentoCMED.tipo_produto)
        )
        tipos_dict = {str(t[0] or "Não informado"): t[1] for t in tipos.all()}

        ultima = await session.exec(
            select(func.max(MedicamentoCMED.data_atualizacao))
        )
        ultima_data = ultima.one()

        return {
            "total_medicamentos": total_count,
            "por_tipo": tipos_dict,
            "ultima_atualizacao": str(ultima_data) if ultima_data else None,
        }
